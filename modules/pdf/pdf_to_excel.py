import os
import tempfile
import aspose.pdf as ap
from openpyxl import load_workbook

def convert_pdf_to_xlsx(input_pdf_path):
    """
    Convert a PDF to Excel and remove the Aspose watermark in one step.
    Returns: path to the cleaned Excel file.
    """
    # Create a temp Excel file for the conversion
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        temp_xlsx_path = tmp_excel.name

    # Convert PDF → Excel
    doc = ap.Document(input_pdf_path)
    save_options = ap.ExcelSaveOptions()
    save_options.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
    save_options.insert_blank_column_at_first = False
    save_options.minimize_the_number_of_worksheets = True
    save_options.uniform_worksheets = False
    doc.save(temp_xlsx_path, save_options)

    # Remove watermark text from all cells
    watermark_text = "Evaluation Only. Created with Aspose.PDF. Copyright 2002-2025 Aspose Pty Ltd."
    wb = load_workbook(temp_xlsx_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and watermark_text in str(cell.value):
                    cell.value = None

    # Save cleaned Excel to a new temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as cleaned_file:
        cleaned_xlsx_path = cleaned_file.name
    wb.save(cleaned_xlsx_path)
    wb.close()

    # Remove the intermediate Excel file
    try:
        os.remove(temp_xlsx_path)
    except FileNotFoundError:
        pass

    return cleaned_xlsx_path

