import os
import sys
import aspose.pdf as ap
from openpyxl import load_workbook

# Ensure we can import from the modules folder
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
sys.path.append(BASE_DIR)

from modules.excel.excel_to_html import excel_to_html_enhanced


def convert_pdf_to_xlsx(input_pdf_path, output_xlsx_path):
    """
    Convert a PDF to Excel and remove the Aspose watermark.
    Saves the cleaned Excel file to `output_xlsx_path`.
    """
    if not os.path.exists(input_pdf_path):
        raise FileNotFoundError(f"❌ PDF not found: {input_pdf_path}")

    # Convert PDF → Excel (Aspose)
    doc = ap.Document(input_pdf_path)
    save_options = ap.ExcelSaveOptions()
    save_options.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
    save_options.insert_blank_column_at_first = False
    save_options.minimize_the_number_of_worksheets = True
    save_options.uniform_worksheets = False
    doc.save(output_xlsx_path, save_options)

    # Remove Aspose watermark text
    watermark_text = (
        "Evaluation Only. Created with Aspose.PDF. "
        "Copyright 2002-2025 Aspose Pty Ltd."
    )
    wb = load_workbook(output_xlsx_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and watermark_text in str(cell.value):
                    cell.value = None
    wb.save(output_xlsx_path)
    wb.close()

    print(f"✅ PDF → Excel conversion complete: {output_xlsx_path}")


def convert_pdf_to_html(input_pdf, output_html):
    temp_xlsx="output.xlsx"
    """
    Full pipeline:
    PDF → Excel → HTML, then clean up temporary Excel file.
    """
    convert_pdf_to_xlsx(input_pdf, temp_xlsx)
    excel_to_html_enhanced(temp_xlsx, output_html)

    # Remove the temporary Excel file
    if os.path.exists(temp_xlsx):
        os.remove(temp_xlsx)
        print(f"🗑️ Deleted temporary file: {temp_xlsx}")

    print(f"✅ PDF → HTML conversion complete: {output_html}")

