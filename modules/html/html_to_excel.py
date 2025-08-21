
# import os
# import re
# import math
# import pandas as pd
# from bs4 import BeautifulSoup, NavigableString, Tag
# import webcolors
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Alignment
# from openpyxl.styles.borders import Border, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.cell.rich_text import CellRichText, TextBlock
# from openpyxl.cell.text import InlineFont

# def html_color_to_openpyxl_argb(html_color):
#     """Convert HTML color names or hex to ARGB for Excel."""
#     if not html_color:
#         return None
#     html_color = html_color.lower().strip()
#     try:
#         if html_color.startswith('#'):
#             hex_val = html_color.lstrip('#')
#         else:
#             hex_val = webcolors.name_to_hex(html_color).lstrip('#')
#         if len(hex_val) == 3:
#             hex_val = "".join([c * 2 for c in hex_val])
#         if len(hex_val) == 6:
#             return 'FF' + hex_val.upper()
#     except ValueError:
#         return None
#     return None

# def parse_style_attr(style_str: str):
#     """Return dict of style properties from an inline style string."""
#     out = {}
#     if not style_str:
#         return out
#     for part in style_str.split(';'):
#         if ':' in part:
#             k, v = part.split(':', 1)
#             out[k.strip().lower()] = v.strip()
#     return out

# def css_color_to_argb(color):
#     """Return FF RRGGBB or None from CSS color string (name, #hex, rgb())."""
#     if not color:
#         return None
#     color = color.strip()
#     m = re.match(r'rgb\(\s*(\d+),\s*(\d+),\s*(\d+)\s*\)', color, re.I)
#     if m:
#         r, g, b = map(int, m.groups())
#         return f"FF{r:02X}{g:02X}{b:02X}"
#     if color.startswith('#'):
#         h = color.lstrip('#')
#         if len(h) == 3:
#             h = ''.join([c*2 for c in h])
#         if len(h) == 6:
#             return "FF" + h.upper()
#     try:
#         hx = webcolors.name_to_hex(color)
#         return "FF" + hx.lstrip('#').upper()
#     except Exception:
#         return None

# def effective_style_for_node(node):
#     """Get effective style for a node, including ancestors."""
#     style = {"bold": False, "italic": False, "underline": False, "color": None, "font_size": None}
#     cur = node if isinstance(node, Tag) else node.parent
#     ancestors = []
#     while cur:
#         ancestors.append(cur)
#         cur = cur.parent
#     for anc in reversed(ancestors):
#         if not isinstance(anc, Tag):
#             continue
#         t = anc.name.lower()
#         if t in ("b", "strong"):
#             style["bold"] = True
#         if t in ("i", "em"):
#             style["italic"] = True
#         if t == "u":
#             style["underline"] = True
#         if t in ("font", "td", "b", "span") and anc.get("color") and style["color"] is None:
#             argbc = css_color_to_argb(anc.get("color"))
#             if argbc:
#                 style["color"] = argbc
#         s = parse_style_attr(anc.get("style", "") or "")
#         if "color" in s and style["color"] is None:
#             argbc = css_color_to_argb(s["color"])
#             if argbc:
#                 style["color"] = argbc
#         if "font-weight" in s and not style["bold"]:
#             if s["font-weight"].strip().lower() in ("bold", "700"):
#                 style["bold"] = True
#         if "font-style" in s and not style["italic"]:
#             if s["font-style"].strip().lower() == "italic":
#                 style["italic"] = True
#         if "text-decoration" in s and not style["underline"]:
#             if "underline" in s["text-decoration"]:
#                 style["underline"] = True
#         if "font-size" in s and style["font_size"] is None:
#             m = re.search(r'([\d.]+)px', s["font-size"])
#             if m:
#                 style["font_size"] = int(round(float(m.group(1)) * 0.75))
#     return style

# def collect_runs_from_cell(cell_tag: Tag):
#     """Walk the cell content and return a list of (text, style_dict) runs in order."""
#     runs = []
#     def walk(node):
#         if isinstance(node, NavigableString):
#             txt = str(node)
#             if txt.strip():
#                 runs.append((txt, effective_style_for_node(node)))
#         elif isinstance(node, Tag):
#             if node.name.lower() == "br":
#                 runs.append(("\n", effective_style_for_node(node)))
#             else:
#                 if node.name.lower() in ("b", "span", "font"):
#                     text = node.get_text(separator="", strip=False)
#                     if text.strip().startswith('•'):
#                         runs.append((text.strip(), effective_style_for_node(node)))
#                     else:
#                         for child in node.children:
#                             walk(child)
#                 else:
#                     for child in node.children:
#                         walk(child)
#     walk(cell_tag)
#     merged = []
#     for text, st in runs:
#         key = (st["bold"], st["italic"], st["underline"], st["color"], st["font_size"])
#         if merged and merged[-1][1] == key and text != "\n" and merged[-1][0] != "\n":
#             merged[-1] = (merged[-1][0] + text, key)
#         else:
#             merged.append((text, key))
#     out = []
#     for text, key in merged:
#         if text.strip() or text == "\n":
#             out.append((text, {
#                 "bold": key[0], "italic": key[1], "underline": key[2],
#                 "color": key[3], "font_size": key[4]
#             }))
#     return out

# def get_cell_background_color(cell, style_str):
#     """Extract background color from <td> or <th> using style attribute first, then bgcolor."""
#     s = parse_style_attr(style_str)
#     bg_color_html = s.get('background-color')
#     if not bg_color_html:
#         bg_color_html = cell.get('bgcolor')
#     return bg_color_html


# from openpyxl.cell.rich_text import CellRichText
# from openpyxl.styles import Border, Side

# def is_price_or_number(value):
#     """Check if the value is a number or price-like string."""
#     if value is None:
#         return False
#     if isinstance(value, (int, float)):
#         return True
#     try:
#         float(str(value).replace('$', '').replace(',', ''))
#         return True
#     except (ValueError, TypeError):
#         return False

# def is_cell_in_merged_range(ws, row, col):
#     """Check if a cell at (row, col) is part of a merged range."""
#     for merged_range in ws.merged_cells.ranges:
#         min_row = merged_range.min_row
#         max_row = merged_range.max_row
#         min_col = merged_range.min_col
#         max_col = merged_range.max_col
#         if min_row <= row <= max_row and min_col <= col <= max_col:
#             return True
#     return False

# def merge_consecutive_duplicates_in_worksheet(ws, columns_to_check=None):
#     """Merge consecutive identical values in column A, merge other columns if uniform within those groups, and set invisible horizontal borders for non-merged cells."""
#     crucial_column = 1  # Column A is the source for merging

#     # Set default columns_to_check to all columns except A
#     if columns_to_check is None:
#         columns_to_check = [c for c in range(2, ws.max_column + 1)]

#     # Validate columns_to_check
#     for col in columns_to_check:
#         if not isinstance(col, int) or col < 1 or col > ws.max_column:
#             raise ValueError(f"Invalid column index {col}. Must be an integer between 1 and {ws.max_column}.")

#     # Define border style: invisible horizontal (top/bottom), visible vertical (left/right)
#     visible_border = Side(style='thin')
#     no_border = Side(style=None)
#     custom_border = Border(left=visible_border, right=visible_border, top=no_border, bottom=no_border)

#     # Step 1: Find and merge consecutive duplicates in column A
#     groups = []
#     prev_value = None
#     start_row = None

#     for row in range(1, ws.max_row + 1):
#         cell = ws.cell(row=row, column=crucial_column)
#         value = cell.value
#         if isinstance(value, CellRichText):
#             value = str(value)

#         if value == prev_value and value is not None:
#             if not is_price_or_number(value):  # Only merge non-numeric/non-price values
#                 if start_row is None:
#                     start_row = row - 1
#         else:
#             if start_row is not None:
#                 end = row - 1
#                 num_rows = end - start_row + 1
#                 if num_rows > 1:
#                     ws.merge_cells(start_row=start_row, start_column=crucial_column,
#                                    end_row=end, end_column=crucial_column)
#                     groups.append((start_row, end))
#                 start_row = None

#         prev_value = value

#     # Handle the last group in column A
#     if start_row is not None:
#         end = ws.max_row
#         num_rows = end - start_row + 1
#         if num_rows > 1:
#             ws.merge_cells(start_row=start_row, start_column=crucial_column,
#                            end_row=end, end_column=crucial_column)
#             groups.append((start_row, end))

#     # Step 2: For each merged group in column A, check other columns and merge if values are identical
#     # Apply invisible horizontal borders to non-merged cells
#     for start, end in groups:
#         num_rows = end - start + 1
#         if num_rows <= 1:
#             continue  # No merge possible for single rows

#         for col in columns_to_check:
#             values = []
#             for r in range(start, end + 1):
#                 cell = ws.cell(row=r, column=col)
#                 v = cell.value
#                 if isinstance(v, CellRichText):
#                     v = str(v)
#                 values.append(v)

#             if values:
#                 first_v = values[0]
#                 if first_v is not None and all(v == first_v for v in values) and not is_price_or_number(first_v):
#                     # Merge cells if identical and not price/number
#                     ws.merge_cells(start_row=start, start_column=col,
#                                    end_row=end, end_column=col)
#                 else:
#                     # Apply invisible horizontal borders to non-merged cells
#                     for r in range(start, end + 1):
#                         if not is_cell_in_merged_range(ws, r, col):
#                             cell = ws.cell(row=r, column=col)
#                             cell.border = custom_border

# def convert_html_to_excel(input_file, output_file, columns_to_merge=None):
#     PIXELS_TO_EXCEL_UNITS = 8.43
#     with open(input_file, 'r', encoding='utf-8') as f:
#         html_content = f.read()
#     soup = BeautifulSoup(html_content, 'html.parser')
#     tables = soup.find_all('table')
#     if not tables:
#         text = soup.get_text(separator='\n', strip=True)
#         df = pd.DataFrame([line for line in text.split('\n') if line], columns=['Content'])
#         df.to_excel(output_file, index=False)
#         return
#     workbook = Workbook()
#     worksheet = workbook.active
#     thin_black_side = Side(style='thin', color='FF000000')
#     default_border = Border(left=thin_black_side, right=thin_black_side,
#                             top=thin_black_side, bottom=thin_black_side)
#     master_layout_pixels = []
#     max_cols = 0
#     for table in tables:
#         cols = table.find_all('col')
#         if len(cols) > max_cols:
#             max_cols = len(cols)
#             master_layout_pixels = []
#             for col in cols:
#                 style = col.get('style', '')
#                 match = re.search(r'width:\s*(\d+)', style)
#                 if match:
#                     master_layout_pixels.append(int(match.group(1)))
#     if not master_layout_pixels:
#         pd.read_html(html_content)[0].to_excel(output_file, index=False)
#         return
#     master_layout_excel_units = [px / PIXELS_TO_EXCEL_UNITS for px in master_layout_pixels]
#     for i, width in enumerate(master_layout_excel_units):
#         worksheet.column_dimensions[get_column_letter(i + 1)].width = width
#     current_row_excel = 1
#     skip_positions = set()
#     for table_idx, table in enumerate(tables):
#         local_layout_pixels = []
#         local_cols = table.find_all('col')
#         if local_cols:
#             for col in local_cols:
#                 style = col.get('style', '')
#                 match = re.search(r'width:\s*(\d+)', style)
#                 if match:
#                     local_layout_pixels.append(int(match.group(1)))
#         if table_idx > 0:
#             current_row_excel += 1
#         rows = table.find_all('tr')
#         for row in rows:
#             cells = row.find_all(['td', 'th'])
#             cell_bullet_splits = []
#             cell_runs = []
#             max_bullet_count = 0
#             for cell_idx, cell in enumerate(cells):
#                 runs = collect_runs_from_cell(cell)
#                 cell_runs.append(runs)
#                 bullet_items = []
#                 current_item = ''
#                 for run_text, run_style in runs:
#                     if run_text == "\n":
#                         if current_item.strip().startswith('•'):
#                             bullet_items.append(current_item.strip())
#                         current_item = ''
#                     else:
#                         current_item += run_text
#                 if current_item.strip().startswith('•'):
#                     bullet_items.append(current_item.strip())
#                 elif current_item.strip():
#                     bullet_items.append(current_item.strip())
#                 if not bullet_items:
#                     bullet_items = [cell.get_text(strip=True)] if cell.get_text(strip=True) else [""]
#                 cell_bullet_splits.append(bullet_items)
#                 max_bullet_count = max(max_bullet_count, len(bullet_items))
#             for cell_idx in range(len(cell_bullet_splits)):
#                 if len(cell_bullet_splits[cell_idx]) < max_bullet_count and len(cell_bullet_splits[cell_idx]) == 1 and not cell_bullet_splits[cell_idx][0].startswith('•'):
#                     cell_bullet_splits[cell_idx] = cell_bullet_splits[cell_idx] * max_bullet_count
#                 elif len(cell_bullet_splits[cell_idx]) < max_bullet_count:
#                     cell_bullet_splits[cell_idx].extend([""] * (max_bullet_count - len(cell_bullet_splits[cell_idx])))
#             for bullet_row in range(max_bullet_count):
#                 current_col_excel = 1
#                 for cell_idx, cell in enumerate(cells):
#                     while (current_row_excel + bullet_row, current_col_excel) in skip_positions:
#                         current_col_excel += 1
#                     style_str = cell.get('style', '') + row.get('style', '')
#                     bg_color_html = get_cell_background_color(cell, style_str)
#                     align_map = {'center': 'center', 'left': 'left', 'right': 'right', 'justify': 'justify'}
#                     text_align = 'general'
#                     align_match = re.search(r'text-align:\s*([^;]+)', style_str, re.IGNORECASE)
#                     if align_match:
#                         text_align = align_map.get(align_match.group(1).strip().lower(), 'general')
#                     is_bold = bool(cell.find('b') or cell.find('strong') or cell.name == 'th' or 'font-weight: bold' in style_str)
#                     is_italic = 'font-style: italic' in style_str or cell.find('i')
#                     is_underline = 'text-decoration: underline' in style_str or bool(cell.find('u'))
#                     is_strike = 'text-decoration: line-through' in style_str
#                     font_family = None
#                     font_family_match = re.search(r'font-family:\s*([^;]+)', style_str, re.IGNORECASE)
#                     if font_family_match:
#                         font_family = font_family_match.group(1).split(',')[0].strip().strip("'\"")
#                     font_size = None
#                     font_size_match = re.search(r'font-size:\s*([\d.]+)px', style_str, re.IGNORECASE)
#                     if font_size_match:
#                         font_size = int(round(float(font_size_match.group(1)) * 0.75))
#                     html_colspan = int(cell.get('colspan', 1))
#                     html_rowspan = int(cell.get('rowspan', 1))
#                     target_pixel_width = 0
#                     if local_layout_pixels and cell_idx < len(local_layout_pixels):
#                         for i in range(html_colspan):
#                             if (cell_idx + i) < len(local_layout_pixels):
#                                 target_pixel_width += local_layout_pixels[cell_idx + i]
#                     excel_colspan = 0
#                     covered_width = 0
#                     if target_pixel_width > 0:
#                         start_master_col_idx = current_col_excel - 1
#                         while covered_width < (target_pixel_width * 0.9) and \
#                               (start_master_col_idx + excel_colspan) < len(master_layout_pixels):
#                             covered_width += master_layout_pixels[start_master_col_idx + excel_colspan]
#                             excel_colspan += 1
#                     excel_colspan = max(1, excel_colspan)
#                     bullet_items = cell_bullet_splits[cell_idx]
#                     text = bullet_items[bullet_row]
#                     runs = []
#                     matched = False
#                     for run_text, run_style in cell_runs[cell_idx]:
#                         clean_run_text = run_text.strip()
#                         clean_text = text.strip()
#                         clean_bullet_text = clean_text[1:].strip() if clean_text.startswith('•') else clean_text
#                         clean_run_bullet_text = clean_run_text[1:].strip() if clean_run_text.startswith('•') else clean_run_text
#                         if clean_run_text == clean_text or clean_run_bullet_text == clean_bullet_text or \
#                            (clean_text == "" and clean_run_text == ""):
#                             runs.append((text, run_style))
#                             matched = True
#                             break
#                     if not matched:
#                         cell_style = effective_style_for_node(cell)
#                         runs.append((text, cell_style))
#                     alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)
#                     fill = None
#                     bg_color_argb = html_color_to_openpyxl_argb(bg_color_html)
#                     if bg_color_argb:
#                         try:
#                             fill = PatternFill(start_color=bg_color_argb, end_color=bg_color_argb, fill_type="solid")
#                         except ValueError:
#                             fill = None
#                     excel_cell = worksheet.cell(row=current_row_excel + bullet_row, column=current_col_excel)
#                     excel_cell.alignment = alignment
#                     excel_cell.border = default_border
#                     if fill:
#                         excel_cell.fill = fill
#                     if runs:
#                         rich_text = CellRichText()
#                         for run_text, run_style in runs:
#                             if run_text.strip() or run_text == "\n":
#                                 inline_font = InlineFont(
#                                     b=run_style["bold"],
#                                     i=run_style["italic"],
#                                     u='single' if run_style["underline"] else None,
#                                     color=run_style["color"],
#                                     sz=run_style["font_size"]
#                                 )
#                                 rich_text.append(TextBlock(inline_font, run_text))
#                         if rich_text:
#                             excel_cell.value = rich_text
#                     else:
#                         if text.strip():
#                             excel_cell.value = text
#                         font = Font(
#                             name=font_family if font_family else None,
#                             size=font_size if font_size else None,
#                             bold=bool(is_bold),
#                             italic=bool(is_italic),
#                             underline='single' if is_underline else None,
#                             strike=bool(is_strike)
#                         )
#                         excel_cell.font = font
#                     if excel_colspan > 1 or html_rowspan > 1:
#                         end_col = current_col_excel + excel_colspan - 1
#                         end_row = current_row_excel + bullet_row + html_rowspan - 1
#                         worksheet.merge_cells(start_row=current_row_excel + bullet_row, start_column=current_col_excel,
#                                              end_row=end_row, end_column=end_col)
#                         for r in range(current_row_excel + bullet_row, end_row + 1):
#                             for c in range(current_col_excel, end_col + 1):
#                                 if not (r == current_row_excel + bullet_row and c == current_col_excel):
#                                     skip_positions.add((r, c))
#                                 worksheet.cell(row=r, column=c).border = default_border
#                     current_col_excel += excel_colspan
#             current_row_excel += max_bullet_count
#     POINTS_PER_LINE = 15.0
#     for row_index in range(1, worksheet.max_row + 1):
#         max_lines_in_row = 1
#         for cell in worksheet[row_index]:
#             if not cell.value:
#                 continue
#             effective_width_units = worksheet.column_dimensions[cell.column_letter].width
#             text = str(cell.value)
#             lines_from_newlines = text.count('\n') + 1
#             lines_from_wrapping = math.ceil(len(text) / (effective_width_units / 1.1)) if effective_width_units else 1
#             cell_lines = max(lines_from_newlines, lines_from_wrapping)
#             max_lines_in_row = max(max_lines_in_row, cell_lines)
#         worksheet.row_dimensions[row_index].height = max_lines_in_row * POINTS_PER_LINE
#     merge_consecutive_duplicates_in_worksheet(worksheet, columns_to_merge)
#     workbook.save(output_file)
#     print(f"Conversion and merging done. Saved to {output_file}")

# # convert_html_to_excel('output.html','output.xlsx')


import os
import re
import math
import pandas as pd
from bs4 import BeautifulSoup, NavigableString, Tag
import webcolors
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

def clean_text(text):
    """Remove newlines and reduce multiple spaces to a single space."""
    cleaned = re.sub(r'\s+', ' ', text.strip())
    return cleaned

def add_missing_semicolons(html_content):
    """Add missing semicolons to style attributes in HTML content."""
    style_pattern = r'style=[\'"]([^\'"]*)[\'"]'

    def replace_styles(match):
        style_content = match.group(1)
        properties = [prop.strip() for prop in style_content.split(';') if prop.strip()]
        updated_properties = []
        for prop in properties:
            if ':' in prop and not prop.endswith(';'):
                prop += ';'
            updated_properties.append(prop)
        new_style_content = ' '.join(updated_properties)
        return f'style="{new_style_content}"'

    return re.sub(style_pattern, replace_styles, html_content)

def html_color_to_openpyxl_argb(html_color):
    """Convert HTML color names or hex to ARGB for Excel."""
    if not html_color:
        return None
    html_color = html_color.lower().strip()
    try:
        if html_color.startswith('#'):
            hex_val = html_color.lstrip('#')
        else:
            hex_val = webcolors.name_to_hex(html_color).lstrip('#')
        if len(hex_val) == 3:
            hex_val = "".join([c * 2 for c in hex_val])
        if len(hex_val) == 6:
            return 'FF' + hex_val.upper()
    except ValueError:
        return None
    return None

def parse_style_attr(style_str: str):
    """Return dict of style properties from an inline style string."""
    out = {}
    if not style_str:
        return out
    for part in style_str.split(';'):
        if ':' in part:
            k, v = part.split(':', 1)
            out[k.strip().lower()] = v.strip()
    return out

def css_color_to_argb(color):
    """Return FF RRGGBB or None from CSS color string (name, #hex, rgb())."""
    if not color:
        return None
    color = color.strip()
    m = re.match(r'rgb\(\s*(\d+),\s*(\d+),\s*(\d+)\s*\)', color, re.I)
    if m:
        r, g, b = map(int, m.groups())
        return f"FF{r:02X}{g:02X}{b:02X}"
    if color.startswith('#'):
        h = color.lstrip('#')
        if len(h) == 3:
            h = ''.join([c*2 for c in h])
        if len(h) == 6:
            return "FF" + h.upper()
    try:
        hx = webcolors.name_to_hex(color)
        return "FF" + hx.lstrip('#').upper()
    except Exception:
        return None

def effective_style_for_node(node):
    """Get effective style for a node, including ancestors."""
    style = {"bold": False, "italic": False, "underline": False, "color": None, "font_size": None}
    cur = node if isinstance(node, Tag) else node.parent
    ancestors = []
    while cur:
        ancestors.append(cur)
        cur = cur.parent
    for anc in reversed(ancestors):
        if not isinstance(anc, Tag):
            continue
        t = anc.name.lower()
        if t in ("b", "strong"):
            style["bold"] = True
        if t in ("i", "em"):
            style["italic"] = True
        if t == "u":
            style["underline"] = True
        if t in ("font", "td", "b", "span") and anc.get("color") and style["color"] is None:
            argbc = css_color_to_argb(anc.get("color"))
            if argbc:
                style["color"] = argbc
        s = parse_style_attr(anc.get("style", "") or "")
        if "color" in s and style["color"] is None:
            argbc = css_color_to_argb(s["color"])
            if argbc:
                style["color"] = argbc
        if "font-weight" in s and not style["bold"]:
            if s["font-weight"].strip().lower() in ("bold", "700"):
                style["bold"] = True
        if "font-style" in s and not style["italic"]:
            if s["font-style"].strip().lower() == "italic":
                style["italic"] = True
        if "text-decoration" in s and not style["underline"]:
            if "underline" in s["text-decoration"]:
                style["underline"] = True
        if "font-size" in s and style["font_size"] is None:
            m = re.search(r'([\d.]+)px', s["font-size"])
            if m:
                style["font_size"] = int(round(float(m.group(1)) * 0.75))
    return style

def collect_runs_from_cell(cell_tag: Tag):
    """Walk the cell content and return a list of (text, style_dict) runs in order."""
    runs = []
    def walk(node):
        if isinstance(node, NavigableString):
            txt = str(node).replace('\n', '')  # Remove any stray newlines
            if txt.strip():
                runs.append((txt, effective_style_for_node(node)))
        elif isinstance(node, Tag):
            if node.name.lower() == "br":
                runs.append(("\n", effective_style_for_node(node)))
            else:
                for child in node.children:
                    walk(child)
    walk(cell_tag)

    # Split runs containing multiple • into separate runs
    processed_runs = []
    for text, style in runs:
        if text != "\n" and '•' in text:
            items = re.split(r'(?=•)', text)
            for item in items:
                item = item.strip()
                if item:
                    if item.startswith('•'):
                        if processed_runs and processed_runs[-1][0] != "\n":
                            processed_runs.append(("\n", style))
                        processed_runs.append((item, style))
                    else:
                        processed_runs.append((item, style))
        else:
            processed_runs.append((text, style))

    # Merge consecutive runs with same style, respecting newlines
    merged = []
    for text, style in processed_runs:
        key = (style["bold"], style["italic"], style["underline"], style["color"], style["font_size"])
        if merged and merged[-1][1] == key and text != "\n" and merged[-1][0] != "\n":
            merged[-1] = (merged[-1][0] + text, key)
        else:
            merged.append((text, key))

    # Filter out empty runs and ensure newlines are preserved
    out = []
    for text, key in merged:
        if text.strip() or text == "\n":
            out.append((text, {
                "bold": key[0], "italic": key[1], "underline": key[2],
                "color": key[3], "font_size": key[4]
            }))
    return out

def get_cell_background_color(cell, style_str):
    """Extract background color from <td> or <th> using style attribute first, then bgcolor."""
    s = parse_style_attr(style_str)
    bg_color_html = s.get('background-color')
    if not bg_color_html:
        bg_color_html = cell.get('bgcolor')
    return bg_color_html

def merge_consecutive_duplicates_in_worksheet(ws, max_bullet_count, row_start_indices):
    """Merge cells based on bullet count for specific columns, applying borders."""
    thin = Side(style="thin", color="FF000000")
    medium = Side(style="thin", color="FF000000")
    no_border = Border(left=None, right=None, top=None, bottom=None)

    # Clear existing borders
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = no_border

    # Find the row after "BASIC INFO"
    basic_info_row = None
    for r in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(r, 1).value or '').strip().upper()
        if cell_value == "BASIC INFO":
            basic_info_row = r
            break
    apply_start = basic_info_row + 1 if basic_info_row is not None else 1

    # Create row_to_group mapping
    row_to_group = {}
    for start_row, bullet_count in row_start_indices:
        for r in range(start_row, start_row + bullet_count):
            row_to_group[r] = start_row

    # Merge cells based on bullet count, including additional columns if applicable
    for start_row, bullet_count in row_start_indices:
        if bullet_count > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + bullet_count - 1, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row + bullet_count - 1, end_column=4)
            # Center align merged cells
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=start_row, column=4).alignment = Alignment(horizontal='center', vertical='center')

            # Merge additional columns (5+) 
            for col in range(5, ws.max_column + 1):
                first_cell = ws.cell(start_row, col)
                all_blank_below = all(
                    ws.cell(r, col).value is None or str(ws.cell(r, col).value).strip() == ''
                    for r in range(start_row + 1, start_row + bullet_count)
                )
                all_blank = (first_cell.value is None or str(first_cell.value).strip() == '') and all_blank_below

                if first_cell.value is not None and str(first_cell.value).strip() and all_blank_below:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + bullet_count - 1, end_column=col)
                    # Set vertical alignment to center
                    first_cell.alignment = Alignment(horizontal=first_cell.alignment.horizontal, vertical='center', wrap_text=True)
                elif start_row >= apply_start and all_blank:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + bullet_count - 1, end_column=col)
                    # Ensure value is None and set alignment if needed
                    first_cell.value = None
                    first_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Apply borders
    status_col = 4
    for row in range(1, ws.max_row + 1):
        group_start = row_to_group.get(row, row)
        status_cell = ws.cell(group_start, status_col)
        status_value = ''
        if status_cell.value is not None:
            if isinstance(status_cell.value, CellRichText):
                status_value = str(status_cell.value)
            else:
                status_value = str(status_cell.value)
        is_no = "NO" in status_value.strip().upper()
        border_style = medium if is_no else thin

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            is_merged = False
            merged_range = None
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    is_merged = True
                    merged_range = mr
                    break

            if is_merged:
                # Apply border to outer edges of merged range
                left = border_style if col == merged_range.min_col else None
                right = border_style if col == merged_range.max_col else None
                top = border_style if row == merged_range.min_row else None
                bottom = border_style if row == merged_range.max_row else None
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            else:
                # Apply full border to non-merged cells
                cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Adjust horizontal borders for columns 2 and 3 in multi-bullet groups to remove internal horizontals
    for start_row, bullet_count in row_start_indices:
        if bullet_count > 1:
            for sub_row in range(start_row, start_row + bullet_count):
                is_first = sub_row == start_row
                is_last = sub_row == start_row + bullet_count - 1
                group_start = row_to_group[sub_row]
                status_value = str(ws.cell(group_start, status_col).value or '')
                is_no = "NO" in status_value.strip().upper()
                border_style = medium if is_no else thin
                for col in [2, 3]:
                    cell = ws.cell(sub_row, col)
                    border = cell.border
                    top_side = border_style if is_first else None
                    bottom_side = border_style if is_last else None
                    new_border = Border(left=border.left, right=border.right, top=top_side, bottom=bottom_side)
                    cell.border = new_border

def convert_html_to_excel(input_file, output_file, columns_to_merge=None):
    PIXELS_TO_EXCEL_UNITS = 8.43
    with open(input_file, 'r', encoding='utf-8') as f:
        html_content = f.read()

    html_content = add_missing_semicolons(html_content)
    html_content = re.sub(r'<td([^>]*)>(.*?)</td>',
                         lambda m: f'<td{m.group(1)}>{clean_text(m.group(2))}</td>',
                         html_content,
                         flags=re.DOTALL)

    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')
    if not tables:
        text = soup.get_text(separator='\n', strip=True)
        df = pd.DataFrame([line for line in text.split('\n') if line], columns=['Content'])
        df.to_excel(output_file, index=False)
        return
    workbook = Workbook()
    worksheet = workbook.active
    thin_black_side = Side(style='thin', color='FF000000')
    master_layout_pixels = []
    max_cols = 0
    for table in tables:
        cols = table.find_all('col')
        if len(cols) > max_cols:
            max_cols = len(cols)
            master_layout_pixels = []
            for col in cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match:
                    master_layout_pixels.append(int(match.group(1)))
    if not master_layout_pixels:
        pd.read_html(html_content)[0].to_excel(output_file, index=False)
        return
    master_layout_excel_units = [px / PIXELS_TO_EXCEL_UNITS for px in master_layout_pixels]
    for i, width in enumerate(master_layout_excel_units, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width if i <= 7 else 20
    current_row_excel = 1
    skip_positions = set()
    row_start_indices = []  # Track start row and bullet count for merging
    for table_idx, table in enumerate(tables):
        if table_idx > 0:
            current_row_excel += 1
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            cell_bullet_splits = []
            max_bullet_count = 0
            for cell_idx, cell in enumerate(cells):
                runs = collect_runs_from_cell(cell)
                bullet_groups = []
                current_group = []
                for run_text, run_style in runs:
                    if run_text == "\n":
                        if current_group:
                            bullet_groups.append(current_group)
                            current_group = []
                    else:
                        current_group.append((run_text, run_style))
                if current_group:
                    bullet_groups.append(current_group)
                if not bullet_groups:
                    text = cell.get_text(strip=True)
                    bullet_groups.append([(text, effective_style_for_node(cell))] if text else [("", effective_style_for_node(cell))])
                cell_bullet_splits.append(bullet_groups)
                max_bullet_count = max(max_bullet_count, len(bullet_groups))
            row_start_indices.append((current_row_excel, max_bullet_count))
            for bullet_idx in range(max_bullet_count):
                current_col_excel = 1
                for cell_idx, cell in enumerate(cells):
                    while (current_row_excel + bullet_idx, current_col_excel) in skip_positions:
                        current_col_excel += 1
                    style_str = cell.get('style', '') + row.get('style', '')
                    bg_color_html = get_cell_background_color(cell, style_str)
                    align_map = {'center': 'center', 'left': 'left', 'right': 'right', 'justify': 'justify'}
                    text_align = 'center'
                    align_match = re.search(r'text-align:\s*([^;]+)', style_str, re.IGNORECASE)
                    if align_match:
                        text_align = align_map.get(align_match.group(1).strip().lower(), 'center')
                    html_colspan = int(cell.get('colspan', 1))
                    html_rowspan = int(cell.get('rowspan', 1))
                    target_pixel_width = 0
                    if cell_idx < len(master_layout_pixels):
                        for i in range(html_colspan):
                            if (cell_idx + i) < len(master_layout_pixels):
                                target_pixel_width += master_layout_pixels[cell_idx + i]
                    excel_colspan = 0
                    covered_width = 0
                    if target_pixel_width > 0:
                        start_master_col_idx = current_col_excel - 1
                        while covered_width < (target_pixel_width * 0.9) and (start_master_col_idx + excel_colspan) < len(master_layout_pixels):
                            covered_width += master_layout_pixels[start_master_col_idx + excel_colspan]
                            excel_colspan += 1
                    excel_colspan = max(1, excel_colspan)
                    run_list = cell_bullet_splits[cell_idx][bullet_idx] if bullet_idx < len(cell_bullet_splits[cell_idx]) else []
                    alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)
                    fill = None
                    bg_color_argb = html_color_to_openpyxl_argb(bg_color_html)
                    if bg_color_argb:
                        try:
                            fill = PatternFill(start_color=bg_color_argb, end_color=bg_color_argb, fill_type="solid")
                        except ValueError:
                            fill = None
                    excel_cell = worksheet.cell(row=current_row_excel + bullet_idx, column=current_col_excel)
                    # Set left alignment for column A (column 1)
                    if current_col_excel == 1:
                        excel_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        excel_cell.alignment = alignment
                    if fill:
                        excel_cell.fill = fill
                    cell_style = effective_style_for_node(cell)
                    if run_list:
                        rich_text = CellRichText()
                        for run_text, run_style in run_list:
                            if run_text.strip() or run_text == "\n":
                                inline_font = InlineFont(
                                    b=run_style["bold"] or (cell_idx == 0),
                                    i=run_style["italic"],
                                    u='single' if run_style["underline"] else None,
                                    color=run_style["color"],
                                    sz=run_style["font_size"]
                                )
                                rich_text.append(TextBlock(inline_font, run_text.strip()))
                        if rich_text:
                            excel_cell.value = rich_text
                    else:
                        text = cell.get_text(strip=True)
                        if text.strip() and bullet_idx == 0:
                            excel_cell.value = text
                        font = Font(
                            size=cell_style["font_size"] if cell_style["font_size"] else None,
                            bold=cell_idx == 0,
                            italic=cell_style["italic"],
                            underline='single' if cell_style["underline"] else None
                        )
                        excel_cell.font = font
                    if excel_colspan > 1 or html_rowspan > 1:
                        end_col = current_col_excel + excel_colspan - 1
                        end_row = current_row_excel + bullet_idx + html_rowspan - 1
                        worksheet.merge_cells(start_row=current_row_excel + bullet_idx, start_column=current_col_excel,
                                             end_row=end_row, end_column=end_col)
                        for r in range(current_row_excel + bullet_idx, end_row + 1):
                            for c in range(current_col_excel, end_col + 1):
                                if not (r == current_row_excel + bullet_idx and c == current_col_excel):
                                    skip_positions.add((r, c))
                    current_col_excel += excel_colspan
            current_row_excel += max_bullet_count
    POINTS_PER_LINE = 15.0
    for row_index in range(1, worksheet.max_row + 1):
        max_lines_in_row = 1
        for cell in worksheet[row_index]:
            if not cell.value:
                continue
            effective_width_units = worksheet.column_dimensions[cell.column_letter].width
            text = str(cell.value)
            lines_from_newlines = text.count('\n') + 1
            lines_from_wrapping = math.ceil(len(text) / (effective_width_units / 1.1)) if effective_width_units else 1
            cell_lines = max(lines_from_newlines, lines_from_wrapping)
            max_lines_in_row = max(max_lines_in_row, cell_lines)
        worksheet.row_dimensions[row_index].height = max_lines_in_row * POINTS_PER_LINE
    merge_consecutive_duplicates_in_worksheet(worksheet, max_bullet_count, row_start_indices)
    workbook.save(output_file)
    print(f"Conversion and merging done. Saved to {output_file}")

# Example usage
# convert_html_to_excel('Html Export (2).html', 'output.xlsx')