import os
import sys
import pythoncom
from docx2pdf import convert
import aspose.pdf as ap
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET

# Ensure we can import from the modules folder
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
sys.path.append(BASE_DIR)


# Note: The excel_to_html_enhanced function should be in modules.excel.excel_to_html
# For this code, we include it directly as provided
def rgb_to_hex(rgb_color):
    """Convert RGB color to hex format"""
    if not rgb_color or rgb_color == "00000000":
        return None
    if len(rgb_color) == 8:  # ARGB format
        rgb_color = rgb_color[2:]  # Remove alpha channel
    if len(rgb_color) == 6:
        return f"#{rgb_color}"
    return None


def get_cell_style(cell):
    """Extract styling information from an Excel cell"""
    styles = {}

    # Background color
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        bg_color = rgb_to_hex(cell.fill.start_color.rgb)
        if bg_color and bg_color != "#FFFFFF":  # Skip white background
            styles['background-color'] = bg_color

    # Font color
    if cell.font and cell.font.color and cell.font.color.rgb:
        font_color = rgb_to_hex(cell.font.color.rgb)
        if font_color and font_color != "#000000":  # Skip black text
            styles['color'] = font_color

    # Font weight
    if cell.font and cell.font.bold:
        styles['font-weight'] = 'bold'

    # Font style
    if cell.font and cell.font.italic:
        styles['font-style'] = 'italic'

    # Font size
    if cell.font and cell.font.size:
        styles['font-size'] = f"{cell.font.size}px"

    # Text alignment
    if cell.alignment:
        if cell.alignment.horizontal:
            if cell.alignment.horizontal == 'center':
                styles['text-align'] = 'center'
            elif cell.alignment.horizontal == 'right':
                styles['text-align'] = 'right'
            elif cell.alignment.horizontal == 'left':
                styles['text-align'] = 'left'
        if cell.alignment.vertical:
            if cell.alignment.vertical == 'center':
                styles['vertical-align'] = 'middle'
            elif cell.alignment.vertical == 'top':
                styles['vertical-align'] = 'top'
            elif cell.alignment.vertical == 'bottom':
                styles['vertical-align'] = 'bottom'

    # Borders
    if cell.border:
        border_parts = []
        if cell.border.top and cell.border.top.style:
            border_parts.append('border-top: 1px solid #000')
        if cell.border.bottom and cell.border.bottom.style:
            border_parts.append('border-bottom: 1px solid #000')
        if cell.border.left and cell.border.left.style:
            border_parts.append('border-left: 1px solid #000')
        if cell.border.right and cell.border.right.style:
            border_parts.append('border-right: 1px solid #000')

        if border_parts:
            styles.update(dict(part.split(': ') for part in border_parts))

    return styles


def styles_to_css(styles):
    """Convert styles dictionary to CSS string"""
    if not styles:
        return ""
    css_parts = []
    for prop, value in styles.items():
        css_parts.append(f"{prop}: {value}")
    return "; ".join(css_parts)


def excel_to_html_enhanced(excel_path, html_path):
    """
    Enhanced Excel to HTML converter that preserves exact visual appearance
    """
    try:
        # Load workbook with openpyxl for styling and structure
        wb = load_workbook(excel_path, data_only=False)

        # Start building HTML with Excel-like styling
        html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel to HTML Conversion</title>
    <style>
        :root {
            --base-font-size: 1rem;
            --table-padding: 0.5rem;
            --container-padding: 1rem;
        }

        * {
            box-sizing: border-box;
        }

        body {
            font-family: Calibri, Arial, sans-serif;
            font-size: var(--base-font-size);
            margin: 0;
            padding: var(--container-padding);
            background-color: #f0f0f0;
            width: 100%;
            overflow-x: auto;
        }

        .excel-container {
            background-color: white;
            border: 1px solid #c0c0c0;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            margin: 0 auto;
            width: 100%;
            max-width: 100%;
            overflow-x: auto;
        }

        .sheet-tab {
            background-color: #e0e0e0;
            padding: 0.8rem 1.2rem;
            border-bottom: 1px solid #c0c0c0;
            font-weight: bold;
            color: #333;
            font-size: 1.1rem;
        }

        table {
            border-collapse: collapse;
            width: 100%;
            table-layout: fixed;
            font-family: Calibri, Arial, sans-serif;
            font-size: 100%;
        }

        th, td {
            border: 1px solid #c0c0c0;
            padding: 0.5rem;
            vertical-align: middle;
            word-wrap: break-word;
            overflow-wrap: break-word;
            hyphens: auto;
            font-size: 100%;
            line-height: 1.4;
        }

        th {
            background-color: #f0f0f0;
            font-weight: normal;
            position: sticky;
            top: 0;
        }

        .empty-cell {
            color: transparent;
        }

        .cell-content {
            min-height: 1.5rem;
            display: block;
        }

        /* Responsive text and layout */
        @media screen and (max-width: 1200px) {
            :root {
                --base-font-size: 0.95rem;
            }

            th, td {
                padding: 0.4rem;
                font-size: 0.95em;
            }
        }

        @media screen and (max-width: 992px) {
            :root {
                --base-font-size: 0.9rem;
            }

            .sheet-tab {
                padding: 0.7rem 1rem;
                font-size: 1.05rem;
            }

            th, td {
                padding: 0.35rem;
                font-size: 0.9em;
            }
        }

        @media screen and (max-width: 768px) {
            :root {
                --base-font-size: 0.85rem;
                --container-padding: 0.5rem;
            }

            body {
                padding: 0.5rem;
            }

            .sheet-tab {
                padding: 0.6rem 0.8rem;
                font-size: 1rem;
            }

            th, td {
                padding: 0.3rem;
                font-size: 0.85em;
            }
        }

        @media screen and (max-width: 576px) {
            :root {
                --base-font-size: 0.8rem;
                --container-padding: 0.25rem;
            }

            body {
                padding: 0.25rem;
            }

            .sheet-tab {
                padding: 0.5rem 0.6rem;
                font-size: 0.95rem;
            }

            th, td {
                padding: 0.2rem;
                font-size: 0.8em;
            }
        }
    </style>
</head>
<body>
"""

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            html_content += f'    <div class="excel-container">\n'
            html_content += f'        <div class="sheet-tab">{sheet_name}</div>\n'
            html_content += '        <table>\n'

            # Find the actual data range
            max_row = ws.max_row
            max_col = ws.max_column

            # Process each row
            for row_idx in range(1, max_row + 1):
                html_content += '            <tr>\n'

                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value

                    # Get cell styling
                    styles = get_cell_style(cell)
                    style_attr = f' style="{styles_to_css(styles)}"' if styles else ""

                    # Handle cell content
                    if cell_value is None or cell_value == "":
                        cell_content = '<span class="empty-cell">&nbsp;</span>'
                    else:
                        # Format the value appropriately
                        if isinstance(cell_value, (int, float)):
                            if cell.number_format and '%' in cell.number_format:
                                cell_content = f"{cell_value:.1%}"
                            elif isinstance(cell_value, float) and cell_value.is_integer():
                                cell_content = str(int(cell_value))
                            else:
                                cell_content = str(cell_value)
                        else:
                            cell_content = str(cell_value)

                    # Determine if it's a header cell (first row or bold)
                    is_header = row_idx == 1 or (cell.font and cell.font.bold)
                    cell_tag = 'th' if is_header else 'td'

                    html_content += f'                <{cell_tag}{style_attr}><span class="cell-content">{cell_content}</span></{cell_tag}>\n'

                html_content += '            </tr>\n'

            html_content += '        </table>\n'
            html_content += '    </div>\n'
            html_content += '    <br>\n'

        # Close HTML
        html_content += """</body>
</html>"""

        # Write to file
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        print(f"✅ Enhanced Excel to HTML conversion completed: {html_path}")
        print(f"📊 Processed {len(wb.sheetnames)} sheet(s) with exact visual formatting")

    except Exception as e:
        print(f"❌ Error in enhanced conversion: {e}")
        raise


def convert_word_to_pdf(input_file, output_file):
    """
    Convert a Word document to PDF using docx2pdf.
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"❌ Word file not found: {input_file}")
    if not input_file.lower().endswith(('.docx', '.doc')):
        raise ValueError("Input file must be a Word document (.docx or .doc)")

    pythoncom.CoInitialize()  # Initialize COM for this thread
    try:
        convert(input_file, output_file)
        print(f"✅ Word → PDF conversion complete: {output_file}")
    except Exception as e:
        print(f"❌ Error in Word to PDF conversion: {str(e)}")
    finally:
        pythoncom.CoUninitialize()  # Always clean up


def convert_pdf_to_xlsx(input_pdf_path, output_xlsx_path):
    """
    Convert a PDF to Excel and remove the Aspose watermark.
    Saves the cleaned Excel file to `output_xlsx_path`.
    """
    if not os.path.exists(input_pdf_path):
        raise FileNotFoundError(f"❌ PDF not found: {input_pdf_path}")

    # Convert PDF → Excel (Aspose)
    doc = ap.Document(input_pdf_path)
    save_options = ap.ExcelSaveOptions()
    save_options.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
    save_options.insert_blank_column_at_first = False
    save_options.minimize_the_number_of_worksheets = True
    save_options.uniform_worksheets = False
    doc.save(output_xlsx_path, save_options)

    # Remove Aspose watermark text
    watermark_text = (
        "Evaluation Only. Created with Aspose.PDF. "
        "Copyright 2002-2025 Aspose Pty Ltd."
    )
    wb = load_workbook(output_xlsx_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and watermark_text in str(cell.value):
                    cell.value = None
    wb.save(output_xlsx_path)
    wb.close()

    print(f"✅ PDF → Excel conversion complete: {output_xlsx_path}")


def convert_word_to_html(input_word, output_html):
    """
    Full pipeline: Word → PDF → Excel → HTML, then clean up temporary files.
    """
    temp_pdf = "temp_output.pdf"
    temp_xlsx = "temp_output.xlsx"

    try:
        # Step 1: Convert Word to PDF
        convert_word_to_pdf(input_word, temp_pdf)

        # Step 2: Convert PDF to Excel
        convert_pdf_to_xlsx(temp_pdf, temp_xlsx)

        # Step 3: Convert Excel to HTML
        excel_to_html_enhanced(temp_xlsx, output_html)

        print(f"✅ Word → HTML conversion complete: {output_html}")

    except Exception as e:
        print(f"❌ Error in Word to HTML conversion: {str(e)}")

    finally:
        # Clean up temporary files
        for temp_file in [temp_pdf, temp_xlsx]:
            if os.path.exists(temp_file):
                os.remove(temp_file)
                print(f"🗑️ Deleted temporary file: {temp_file}")

