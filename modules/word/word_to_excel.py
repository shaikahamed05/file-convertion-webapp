import os
import aspose.pdf as ap
from openpyxl import load_workbook
from .word_to_pdf import convert_word_to_pdf

# Example usage:
def convert_word_to_excel(input_file,output_file):
    # Convert Word → PDF
    convert_word_to_pdf(input_file, 'output.pdf')

    def convert_pdf_to_xlsx(input_pdf_path, output_xlsx_path):
        """
        Convert a PDF to Excel and remove the Aspose watermark.
        Saves the cleaned Excel file to `output_xlsx_path` and deletes the PDF.
        """
        # Convert PDF → Excel (Aspose)
        doc = ap.Document(input_pdf_path)
        save_options = ap.ExcelSaveOptions()
        save_options.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
        save_options.insert_blank_column_at_first = False
        save_options.minimize_the_number_of_worksheets = True
        save_options.uniform_worksheets = False
        doc.save(output_xlsx_path, save_options)

        # Remove watermark text
        watermark_text = "Evaluation Only. Created with Aspose.PDF. Copyright 2002-2025 Aspose Pty Ltd."
        wb = load_workbook(output_xlsx_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and watermark_text in str(cell.value):
                        cell.value = None
        wb.save(output_xlsx_path)
        wb.close()

        # Delete the original PDF
        if os.path.exists(input_pdf_path):
            os.remove(input_pdf_path)

        print(f"✅ Conversion complete: {output_xlsx_path}")
    convert_pdf_to_xlsx('output.pdf',output_file)

